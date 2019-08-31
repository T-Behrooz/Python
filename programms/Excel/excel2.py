from openpyxl import Workbook
wb = Workbook()
ws = wb.active
for i in ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','y','w','x','y','z']:
    ws = wb.create_sheet(i)
print(wb.sheetnames)
for i in range(1,101):
    for j in range(1,101):
        ws.cell(row=i, column=j,value=i+j)
wb.save('balances.xlsx')
for sheet in wb:
    print(sheet.title)
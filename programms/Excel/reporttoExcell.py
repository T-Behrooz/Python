import pyodbc

con = pyodbc.connect("DRIVER={SQL Server};SERVER=localhost;DATABASE=test;UID=sa;PWD=14")
do = con.cursor()
row = do.execute("select * from persons")
row1=do.fetchall()
path="D:\\268\\behrooz.xlsx"
from openpyxl import load_workbook
wb = load_workbook(path)
sheet = wb.active
sheet.title = 'گزارش'
sheet['A1']='کد ملی'
sheet['B1']='نام خانوادگی'
sheet['C1']='نام'
sheet['D1']='آدرس'
sheet['E1']='شهر'
# for i in range(2,5):
#     for j,value in enumerate(row1):
#         sheet.cell(row=i,column=1).value = value[0]
#         sheet.cell(row=i,column=2).value = str(value[1])
#         sheet.cell(row=i,column=3).value = str(value[2])
#         sheet.cell(row=i,column=4).value = str(value[3])
#         sheet.cell(row=i,column=5).value = str(value[4])
for i in range(2,5):
    for j,value in enumerate(row1):
       sheet.cell(row=i,column=j).value =str(value[i-2])

wb.save(filename='data.xlsx')
wb.save(path)
do.close()
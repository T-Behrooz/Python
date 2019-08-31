from openpyxl import load_workbook
path="D:\\268\\behrooz.xlsx"
wb=load_workbook(path)
sheet = wb.active
sheet['A1']='this is My programe'
sheet.cell(row = 2, column = 2).value =25
wb.save(path)

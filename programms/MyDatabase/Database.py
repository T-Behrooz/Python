import pyodbc
test='y'
dm = input("آیا می خواهید ورود اطلاعات انجام دهید ؟(y/n)")
if dm=='y':
    count=0
    con=pyodbc.connect("DRIVER={SQL Server};SERVER=localhost;DATABASE=test;UID=sa;PWD=14")
    do=con.cursor()
    while (test=='y'):
       personid=int(input("کد ملي :"))
       lastname=str(input("نام خانوادگي :"))
       firstname=str(input("نام :"))
       address =str(input("آدرس :"))
       city = str(input("شهر :"))
       values=[personid,lastname,firstname,address,city]
       sqlcommand="INSERT INTO Persons values(?,?,?,?,?);"
       do.execute(sqlcommand,values)
       do.commit()
       count+=1
       test=input("ادامه مي دهيد ؟(y/n)")
       do.close()
    print("%d سطر وارد شد" % count)
else :
    dm = input("آیا می خواهید گزارش گیری انجام دهید ؟(y/n)")
    if dm =='y':
        con = pyodbc.connect("DRIVER={SQL Server};SERVER=localhost;DATABASE=test;UID=sa;PWD=14")
        do = con.cursor()
        row=do.execute("select * from persons")
        for i in row :
            print (i)
        print("-----------------------------------------------------")
        t=input("آیا می خواهید پرینتی ار نتایج را داشته باشید ؟ (y/n)")
        if t=='y':
            path=input("مسیر فایل خروجی را وارد کنید:")
           #file = open(path,'w',encoding = 'utf-8')
            row1=do.execute("select * from persons")
           for i in row :
               file.write("-----------------------------------------------------------------------------------------\n")
               file.write(str(i))
               file.write("\n----------------------------------------------------------------------------------------\n")
            file.close()
            path="D:\\268\\behrooz.xlsx"
            from openpyxl import load_workbook
            wb = load_workbook(path)
            sheet = wb.active
            for i in range(1,5):
                for j in range(1,5) :
                    for b in row :
                       sheet.cell(row=i, column=j).value = row1[i-1,j-1]
            wb.save(path)
            do.close()
    else :
       input("برای خروج از برنامه کلیدی را فشار دهید...")

